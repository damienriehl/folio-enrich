"""Tests for the OWL Property (verb/relation) extraction pipeline and components."""

from __future__ import annotations

import json
from unittest.mock import AsyncMock, MagicMock, patch

import pytest

from app.models.annotation import (
    Annotation,
    ConceptMatch,
    PropertyAnnotation,
    Span,
    StageEvent,
)
from app.models.document import (
    CanonicalText,
    DocumentFormat,
    DocumentInput,
    TextChunk,
)
from app.models.job import Job, JobResult, JobStatus


# ── Fixtures ────────────────────────────────────────────────────────────


SAMPLE_LEGAL_TEXT = (
    "The court reversed the grant of summary judgment and remanded "
    "the case for further proceedings. The motion was denied by the judge. "
    "Counsel argued that the statute applied and the contract was drafted "
    "by the defendant. The ruling was affirmed on appeal."
)


def _make_job(text: str = SAMPLE_LEGAL_TEXT) -> Job:
    return Job(
        input=DocumentInput(content=text, format=DocumentFormat.PLAIN_TEXT),
        status=JobStatus.MATCHING,
        result=JobResult(
            canonical_text=CanonicalText(
                full_text=text,
                chunks=[
                    TextChunk(
                        text=text,
                        start_offset=0,
                        end_offset=len(text),
                        chunk_index=0,
                    )
                ],
            ),
            annotations=[
                Annotation(
                    span=Span(start=4, end=9, text="court"),
                    concepts=[
                        ConceptMatch(
                            concept_text="court",
                            folio_iri="https://folio.openlegalstandard.org/Court",
                            folio_label="Court",
                            branches=["Legal Entity"],
                            confidence=0.90,
                            source="reconciled",
                        )
                    ],
                ),
            ],
        ),
    )


# ── Data Model Tests ───────────────────────────────────────────────────


class TestPropertyAnnotationModel:
    def test_property_annotation_creation(self):
        prop = PropertyAnnotation(
            property_text="reversed",
            folio_iri="https://folio.openlegalstandard.org/reversed",
            folio_label="reversed",
            span=Span(start=10, end=18, text="reversed"),
            confidence=0.85,
        )
        assert prop.property_text == "reversed"
        assert prop.folio_label == "reversed"
        assert prop.source == "aho_corasick"
        assert prop.confidence == 0.85
        assert prop.id  # auto-generated

    def test_property_with_domain_range(self):
        prop = PropertyAnnotation(
            property_text="reversed",
            folio_iri="https://folio.openlegalstandard.org/reversed",
            folio_label="reversed",
            span=Span(start=10, end=18, text="reversed"),
            domain_iris=["https://folio.openlegalstandard.org/Court"],
            range_iris=["https://folio.openlegalstandard.org/Decision"],
            confidence=0.90,
        )
        assert len(prop.domain_iris) == 1
        assert len(prop.range_iris) == 1

    def test_property_with_inverse(self):
        prop = PropertyAnnotation(
            property_text="reversed",
            span=Span(start=0, end=8, text="reversed"),
            inverse_of_iri="https://folio.openlegalstandard.org/wasReversedBy",
            inverse_of_label="was reversed by",
        )
        assert prop.inverse_of_iri is not None

    def test_job_result_has_properties(self):
        result = JobResult()
        assert result.properties == []

    def test_job_status_has_extracting_properties(self):
        assert JobStatus.EXTRACTING_PROPERTIES == "extracting_properties"


# ── PropertyMatcher Tests ─────────────────────────────────────────────


class TestPropertyMatcher:
    @pytest.fixture
    def matcher(self):
        from app.services.property.property_matcher import PropertyMatcher
        m = PropertyMatcher()
        m.build()
        return m

    def test_matcher_builds_patterns(self, matcher):
        assert matcher.pattern_count > 0

    def test_matches_multi_word_property(self, matcher):
        text = "The court appeared before the judge."
        results = matcher.match(text)
        # "appeared before" should match as a property (alt label for convened)
        matched_texts = [r.property_text.lower() for r in results]
        assert "appeared before" in matched_texts

    def test_matches_single_word_verb(self, matcher):
        text = "The ruling was affirmed on appeal."
        results = matcher.match(text)
        # "affirmed" should match as a property
        matched_texts = [r.property_text.lower() for r in results]
        assert "affirmed" in matched_texts

    def test_stopwords_excluded(self):
        from app.services.property.property_matcher import _PROPERTY_STOPWORDS
        assert "not" in _PROPERTY_STOPWORDS
        assert "and" in _PROPERTY_STOPWORDS

    def test_word_boundary_enforcement(self, matcher):
        # "denied" should match but not "deniedxx"
        text = "The motion was denied clearly."
        results = matcher.match(text)
        for r in results:
            assert r.property_text == text[r.span.start:r.span.end]

    def test_returns_property_annotations(self, matcher):
        text = "The case was reversed and remanded."
        results = matcher.match(text)
        for r in results:
            assert isinstance(r, PropertyAnnotation)
            assert r.span.start >= 0
            assert r.span.end > r.span.start
            assert r.source == "aho_corasick"
            assert 0 < r.confidence <= 1.0

    def test_empty_text_returns_empty(self, matcher):
        results = matcher.match("")
        assert results == []

    def test_preferred_label_higher_confidence(self, matcher):
        # Preferred labels should have higher base confidence than alt labels
        text = "The case was reversed and remanded."
        results = matcher.match(text)
        for r in results:
            if r.match_type == "preferred":
                assert r.confidence >= 0.85
            elif r.match_type == "alternative":
                assert r.confidence >= 0.75


# ── PropertyDeduplicator Tests ────────────────────────────────────────


class TestPropertyDeduplicator:
    def test_no_overlap_keeps_all(self):
        from app.services.property.property_deduplicator import deduplicate_properties
        props = [
            PropertyAnnotation(
                property_text="reversed",
                span=Span(start=0, end=8, text="reversed"),
                confidence=0.85,
            ),
            PropertyAnnotation(
                property_text="affirmed",
                span=Span(start=20, end=28, text="affirmed"),
                confidence=0.85,
            ),
        ]
        result = deduplicate_properties(props)
        assert len(result) == 2

    def test_contained_property_both_kept(self):
        """Contained property spans should both survive."""
        from app.services.property.property_deduplicator import deduplicate_properties
        props = [
            PropertyAnnotation(
                property_text="appeared",
                span=Span(start=0, end=8, text="appeared"),
                confidence=0.85,
            ),
            PropertyAnnotation(
                property_text="appeared before",
                span=Span(start=0, end=15, text="appeared before"),
                confidence=0.90,
            ),
        ]
        result = deduplicate_properties(props)
        assert len(result) == 2
        texts = {r.property_text for r in result}
        assert "appeared" in texts
        assert "appeared before" in texts

    def test_same_span_higher_confidence_wins(self):
        from app.services.property.property_deduplicator import deduplicate_properties
        props = [
            PropertyAnnotation(
                property_text="denied",
                span=Span(start=0, end=6, text="denied"),
                source="aho_corasick",
                confidence=0.75,
            ),
            PropertyAnnotation(
                property_text="denied",
                span=Span(start=0, end=6, text="denied"),
                source="llm",
                confidence=0.90,
            ),
        ]
        result = deduplicate_properties(props)
        assert len(result) == 1
        assert result[0].confidence == 0.90

    def test_empty_list(self):
        from app.services.property.property_deduplicator import deduplicate_properties
        assert deduplicate_properties([]) == []


# ── LLM Property Identifier Tests ────────────────────────────────────


class TestLLMPropertyIdentifier:
    @pytest.fixture
    def fake_llm(self):
        llm = MagicMock()
        llm.structured = AsyncMock(return_value={
            "properties": [
                {
                    "property_text": "reversed",
                    "folio_label": "reversed",
                    "domain_annotation_ids": [],
                    "range_annotation_ids": [],
                    "confidence": 0.90,
                    "is_new": True,
                },
            ]
        })
        return llm

    async def test_identifies_new_properties(self, fake_llm):
        from app.services.property.llm_property_identifier import LLMPropertyIdentifier
        identifier = LLMPropertyIdentifier(fake_llm)
        chunk = TextChunk(
            text="The court reversed the decision.",
            start_offset=0,
            end_offset=31,
            chunk_index=0,
        )
        results = await identifier.identify_properties(chunk, [], [])
        assert len(results) == 1
        assert results[0].property_text == "reversed"
        assert results[0].source == "llm"

    async def test_handles_llm_failure(self):
        from app.services.property.llm_property_identifier import LLMPropertyIdentifier
        llm = MagicMock()
        llm.structured = AsyncMock(side_effect=Exception("LLM error"))
        identifier = LLMPropertyIdentifier(llm)
        chunk = TextChunk(
            text="Some text.",
            start_offset=0,
            end_offset=10,
            chunk_index=0,
        )
        results = await identifier.identify_properties(chunk, [], [])
        assert results == []

    async def test_batch_processing(self, fake_llm):
        from app.services.property.llm_property_identifier import LLMPropertyIdentifier
        identifier = LLMPropertyIdentifier(fake_llm)
        chunks = [
            TextChunk(text="The court reversed the decision.", start_offset=0, end_offset=31, chunk_index=0),
            TextChunk(text="Another chunk of text.", start_offset=31, end_offset=53, chunk_index=1),
        ]
        results = await identifier.identify_batch(chunks, [], [])
        assert isinstance(results, list)


# ── Pipeline Stage Tests ─────────────────────────────────────────────


class TestPropertyStages:
    async def test_early_stage_runs(self):
        from app.pipeline.stages.property_stage import EarlyPropertyStage
        stage = EarlyPropertyStage()
        assert stage.name == "early_property_extraction"

        job = _make_job()
        result = await stage.execute(job)
        assert result.status == JobStatus.EXTRACTING_PROPERTIES
        assert isinstance(result.result.properties, list)

    async def test_early_stage_finds_properties(self):
        from app.pipeline.stages.property_stage import EarlyPropertyStage
        stage = EarlyPropertyStage()
        job = _make_job()
        result = await stage.execute(job)
        # Should find some properties (reversed, remanded, denied, argued, etc.)
        assert len(result.result.properties) > 0

    async def test_early_stage_skips_when_disabled(self):
        from app.pipeline.stages.property_stage import EarlyPropertyStage
        stage = EarlyPropertyStage()
        job = _make_job()
        with patch("app.config.settings") as mock_settings:
            mock_settings.property_extraction_enabled = False
            result = await stage.execute(job)
        assert len(result.result.properties) == 0

    async def test_early_stage_skips_without_canonical_text(self):
        from app.pipeline.stages.property_stage import EarlyPropertyStage
        stage = EarlyPropertyStage()
        job = Job(
            input=DocumentInput(content="test", format=DocumentFormat.PLAIN_TEXT),
            status=JobStatus.MATCHING,
            result=JobResult(),
        )
        result = await stage.execute(job)
        assert len(result.result.properties) == 0

    async def test_llm_stage_regex_only_mode(self):
        from app.pipeline.stages.property_stage import LLMPropertyStage
        fake_llm = MagicMock()
        fake_llm.structured = AsyncMock()
        stage = LLMPropertyStage(llm=fake_llm)

        job = _make_job()
        with patch("app.config.settings") as mock_settings:
            mock_settings.property_extraction_enabled = True
            mock_settings.property_regex_only = True
            result = await stage.execute(job)

        # LLM should not have been called
        fake_llm.structured.assert_not_called()

    async def test_llm_stage_skips_without_llm(self):
        from app.pipeline.stages.property_stage import LLMPropertyStage
        stage = LLMPropertyStage(llm=None)
        job = _make_job()
        result = await stage.execute(job)
        # Should pass through without error
        assert isinstance(result.result.properties, list)

    async def test_early_stage_logs_activity(self):
        from app.pipeline.stages.property_stage import EarlyPropertyStage
        stage = EarlyPropertyStage()
        job = _make_job()
        result = await stage.execute(job)
        log = result.result.metadata.get("activity_log", [])
        prop_logs = [l for l in log if l.get("stage") == "early_property_extraction"]
        assert len(prop_logs) >= 1


# ── Export Integration Tests ──────────────────────────────────────────


def _make_job_with_properties() -> Job:
    prop1 = PropertyAnnotation(
        property_text="reversed",
        folio_iri="https://folio.openlegalstandard.org/reversed",
        folio_label="reversed",
        folio_definition="The act of reversing a lower court decision",
        span=Span(start=10, end=18, text="reversed"),
        confidence=0.85,
        source="aho_corasick",
        match_type="preferred",
        domain_iris=["https://folio.openlegalstandard.org/Court"],
        range_iris=["https://folio.openlegalstandard.org/Decision"],
    )
    prop2 = PropertyAnnotation(
        property_text="remanded",
        folio_iri="https://folio.openlegalstandard.org/remanded",
        folio_label="remanded",
        span=Span(start=62, end=70, text="remanded"),
        confidence=0.85,
        source="aho_corasick",
        match_type="alternative",
    )

    return Job(
        input=DocumentInput(content=SAMPLE_LEGAL_TEXT, format=DocumentFormat.PLAIN_TEXT),
        status=JobStatus.COMPLETED,
        result=JobResult(
            canonical_text=CanonicalText(
                full_text=SAMPLE_LEGAL_TEXT,
                chunks=[TextChunk(
                    text=SAMPLE_LEGAL_TEXT,
                    start_offset=0,
                    end_offset=len(SAMPLE_LEGAL_TEXT),
                    chunk_index=0,
                )],
            ),
            annotations=[
                Annotation(
                    span=Span(start=4, end=9, text="court"),
                    concepts=[
                        ConceptMatch(
                            concept_text="court",
                            folio_iri="https://folio.openlegalstandard.org/Court",
                            folio_label="Court",
                            branches=["Legal Entity"],
                            confidence=0.95,
                            source="reconciled",
                        )
                    ],
                ),
            ],
            properties=[prop1, prop2],
        ),
    )


class TestExportWithProperties:
    def test_json_export_includes_properties(self):
        from app.services.export.json_exporter import JSONExporter
        job = _make_job_with_properties()
        data = json.loads(JSONExporter().export(job))
        assert "properties" in data
        assert len(data["properties"]) == 2
        assert data["statistics"]["total_properties"] == 2
        assert data["statistics"]["unique_properties"] == 2

    def test_jsonld_export_includes_properties(self):
        from app.services.export.jsonld_exporter import JSONLDExporter
        job = _make_job_with_properties()
        data = json.loads(JSONLDExporter().export(job))
        assert "properties" in data
        assert len(data["properties"]) == 2
        assert data["properties"][0]["@type"] == "owl:ObjectProperty"

    def test_xml_export_includes_properties(self):
        import xml.etree.ElementTree as ET
        from app.services.export.xml_exporter import XMLExporter
        job = _make_job_with_properties()
        root = ET.fromstring(XMLExporter().export(job))
        props = root.find("properties")
        assert props is not None
        assert len(list(props)) == 2

    def test_csv_export_includes_properties(self):
        import csv
        import io
        from app.services.export.csv_exporter import CSVExporter
        job = _make_job_with_properties()
        result = CSVExporter().export(job)
        assert "property_text" in result
        assert "reversed" in result

    def test_jsonl_export_includes_properties(self):
        from app.services.export.jsonl_exporter import JSONLExporter
        job = _make_job_with_properties()
        result = JSONLExporter().export(job)
        lines = [json.loads(l) for l in result.strip().split("\n") if l]
        prop_lines = [l for l in lines if l.get("record_type") == "property"]
        assert len(prop_lines) == 2

    def test_rdf_export_includes_properties(self):
        from app.services.export.rdf_exporter import RDFExporter
        job = _make_job_with_properties()
        result = RDFExporter().export(job)
        assert "ObjectProperty" in result

    def test_brat_export_includes_properties(self):
        from app.services.export.brat_exporter import BratExporter
        job = _make_job_with_properties()
        result = BratExporter().export(job)
        assert "reversed" in result

    def test_html_export_includes_properties(self):
        from app.services.export.html_exporter import HTMLExporter
        job = _make_job_with_properties()
        result = HTMLExporter().export(job)
        assert "folio-property" in result

    def test_elasticsearch_export_includes_properties(self):
        from app.services.export.elasticsearch_exporter import ElasticsearchExporter
        job = _make_job_with_properties()
        result = ElasticsearchExporter().export(job)
        lines = result.strip().split("\n")
        prop_actions = [l for l in lines if "folio-properties" in l]
        assert len(prop_actions) == 2

    def test_neo4j_export_includes_properties(self):
        from app.services.export.neo4j_exporter import Neo4jExporter
        job = _make_job_with_properties()
        result = Neo4jExporter().export(job)
        assert "Property" in result
        assert "CONTAINS_PROPERTY" in result

    def test_excel_export_includes_properties(self):
        from openpyxl import load_workbook
        import io
        from app.services.export.excel_exporter import ExcelExporter
        job = _make_job_with_properties()
        data = ExcelExporter().export(job)
        wb = load_workbook(io.BytesIO(data))
        assert "Properties" in wb.sheetnames
        ws = wb["Properties"]
        assert ws.max_row >= 3  # header + 2 properties


# ── LLM Prompt Tests ────────────────────────────────────────────────


class TestPropertyExtractionPrompt:
    def test_prompt_builds_with_annotations(self):
        from app.services.llm.prompts.property_extraction import (
            build_property_extraction_prompt,
        )
        prompt = build_property_extraction_prompt(
            text="The court reversed the decision.",
            class_annotations=[
                {"id": "ann-1", "label": "Court", "span_text": "court", "branch": "Legal Entity"},
            ],
            existing_properties=[
                {"property_text": "reversed", "folio_label": "reversed", "source": "aho_corasick"},
            ],
            property_labels=["reversed", "affirmed", "denied"],
        )
        assert "reversed" in prompt
        assert "Court" in prompt

    def test_prompt_builds_with_empty_inputs(self):
        from app.services.llm.prompts.property_extraction import (
            build_property_extraction_prompt,
        )
        prompt = build_property_extraction_prompt(
            text="Some text.",
            class_annotations=[],
            existing_properties=[],
            property_labels=[],
        )
        assert "Some text." in prompt
        assert "none found" in prompt.lower()


# ── Config Tests ─────────────────────────────────────────────────────


class TestPropertyConfig:
    def test_default_config_values(self):
        from app.config import Settings
        s = Settings()
        assert s.property_extraction_enabled is True
        assert s.property_regex_only is False
        assert s.llm_property_provider == ""
        assert s.llm_property_model == ""


# ── FolioService Property Tests ──────────────────────────────────────


class TestFolioServiceProperties:
    def test_get_all_property_labels(self):
        from app.services.folio.folio_service import FolioService
        svc = FolioService()
        labels = svc.get_all_property_labels()
        assert len(labels) > 0
        # Check that deprecated properties are excluded
        for key in labels:
            assert "DEPRECATED" not in key.upper()
            assert not key.startswith("zzz:")

    def test_property_labels_have_info(self):
        from app.services.folio.folio_service import FolioService
        svc = FolioService()
        labels = svc.get_all_property_labels()
        # Pick one and verify it has data
        for key, info in labels.items():
            assert info.prop is not None
            assert info.prop.iri
            assert info.label_type in ("preferred", "alternative")
            assert info.matched_label
            break

    def test_prefix_stripping(self):
        from app.services.folio.folio_service import FolioService
        assert FolioService._strip_prefix("folio:reversed") == "reversed"
        assert FolioService._strip_prefix("utbms:Research") == "Research"
        assert FolioService._strip_prefix("affirmed") == "affirmed"

    def test_property_labels_cached(self):
        from app.services.folio.folio_service import FolioService
        svc = FolioService()
        labels1 = svc.get_all_property_labels()
        labels2 = svc.get_all_property_labels()
        assert labels1 is labels2  # Same object from cache
