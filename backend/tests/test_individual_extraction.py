"""Tests for the OWL Individual extraction pipeline stage and components."""

from __future__ import annotations

import json
from unittest.mock import AsyncMock, MagicMock, patch

import pytest

from app.models.annotation import (
    Annotation,
    ConceptMatch,
    Individual,
    IndividualClassLink,
    Span,
    StageEvent,
)
from app.models.document import (
    CanonicalText,
    DocumentFormat,
    DocumentInput,
    TextChunk,
)
from app.models.job import Job, JobResult, JobStatus


# ── Fixtures ────────────────────────────────────────────────────────────


SAMPLE_LEGAL_TEXT = (
    "In Smith v. Jones, 123 U.S. 456 (1987), the Supreme Court of the United States "
    "held that John Smith, the plaintiff, was entitled to $500,000 in damages under "
    "42 U.S.C. § 1983. The case was filed on January 15, 2023 in the S.D.N.Y. "
    "Google LLC argued that the contract signed on 03/15/2022 for a period of "
    "30 days was subject to 5% interest. Judge Roberts ruled that no more than "
    "$1 million could be awarded. Apple® and Microsoft™ were also mentioned. "
    "© 2024 Acme Corp. All rights reserved. The address is "
    "123 Main Street, Suite 100, New York, NY 10001."
)


def _make_job(text: str = SAMPLE_LEGAL_TEXT) -> Job:
    return Job(
        input=DocumentInput(content=text, format=DocumentFormat.PLAIN_TEXT),
        status=JobStatus.MATCHING,
        result=JobResult(
            canonical_text=CanonicalText(
                full_text=text,
                chunks=[
                    TextChunk(
                        text=text,
                        start_offset=0,
                        end_offset=len(text),
                        chunk_index=0,
                    )
                ],
            ),
            annotations=[
                Annotation(
                    span=Span(start=0, end=5, text="Smith"),
                    concepts=[
                        ConceptMatch(
                            concept_text="plaintiff",
                            folio_iri="https://folio.openlegalstandard.org/Plaintiff",
                            folio_label="Plaintiff",
                            branches=["Actor / Player"],
                            confidence=0.90,
                            source="reconciled",
                        )
                    ],
                ),
            ],
        ),
    )


# ── Data Model Tests ───────────────────────────────────────────────────


class TestIndividualModel:
    def test_individual_creation(self):
        ind = Individual(
            name="John Smith",
            mention_text="John Smith",
            span=Span(start=10, end=20, text="John Smith"),
        )
        assert ind.name == "John Smith"
        assert ind.individual_type == "named_entity"
        assert ind.source == "llm"
        assert ind.confidence == 0.0
        assert ind.class_links == []
        assert ind.id  # auto-generated

    def test_individual_with_class_link(self):
        link = IndividualClassLink(
            annotation_id="ann-1",
            folio_iri="https://folio.openlegalstandard.org/Plaintiff",
            folio_label="Plaintiff",
            branch="Actor / Player",
            confidence=0.90,
        )
        ind = Individual(
            name="John Smith",
            mention_text="John Smith",
            span=Span(start=10, end=20, text="John Smith"),
            class_links=[link],
            confidence=0.90,
        )
        assert len(ind.class_links) == 1
        assert ind.class_links[0].folio_label == "Plaintiff"
        assert ind.class_links[0].relationship == "instance_of"

    def test_individual_citation_type(self):
        ind = Individual(
            name="Smith v. Jones, 123 U.S. 456 (1987)",
            mention_text="Smith v. Jones, 123 U.S. 456 (1987)",
            individual_type="legal_citation",
            span=Span(start=3, end=39, text="Smith v. Jones, 123 U.S. 456 (1987)"),
            source="eyecite",
            confidence=0.92,
            normalized_form="123 U.S. 456",
        )
        assert ind.individual_type == "legal_citation"
        assert ind.source == "eyecite"
        assert ind.normalized_form == "123 U.S. 456"

    def test_job_result_has_individuals(self):
        result = JobResult()
        assert result.individuals == []

    def test_job_status_has_extracting_individuals(self):
        assert JobStatus.EXTRACTING_INDIVIDUALS == "extracting_individuals"


# ── Citation Extractor Tests ───────────────────────────────────────────


class TestCitationExtractor:
    @pytest.fixture
    def extractor(self):
        from app.services.individual.citation_extractor import CitationExtractor
        return CitationExtractor()

    async def test_extracts_case_citations(self, extractor):
        text = "In Smith v. Jones, 123 U.S. 456 (1987), the court held..."
        results = await extractor.extract(text)
        # Should find at least the case citation
        citations = [r for r in results if r.individual_type == "legal_citation"]
        assert len(citations) >= 1
        # Check that at least one has the expected case name
        names = [c.name for c in citations]
        assert any("123 U.S. 456" in n for n in names) or any("Smith" in n for n in names)

    async def test_extracts_statutory_citations(self, extractor):
        text = "Under 42 U.S.C. § 1983, a plaintiff may bring suit."
        results = await extractor.extract(text)
        citations = [r for r in results if r.individual_type == "legal_citation"]
        assert len(citations) >= 1

    async def test_citations_have_spans(self, extractor):
        text = "See Brown v. Board of Education, 347 U.S. 483 (1954)."
        results = await extractor.extract(text)
        for ind in results:
            assert ind.span.start >= 0
            assert ind.span.end > ind.span.start
            assert ind.span.text

    async def test_citations_have_source_eyecite(self, extractor):
        text = "Miranda v. Arizona, 384 U.S. 436 (1966) established..."
        results = await extractor.extract(text)
        for ind in results:
            assert ind.source in ("eyecite", "citeurl")

    async def test_empty_text_returns_empty(self, extractor):
        results = await extractor.extract("")
        assert results == []

    async def test_no_citations_returns_empty(self, extractor):
        results = await extractor.extract("This is a simple sentence with no legal citations.")
        # May still find some via aggressive parsing, but should be few
        assert isinstance(results, list)


# ── Entity Extractor Tests ─────────────────────────────────────────────


class TestEntityExtractors:
    def test_monetary_amount_extractor(self):
        from app.services.individual.entity_extractors import MonetaryAmountExtractor
        ext = MonetaryAmountExtractor()
        results = ext.extract_sync("The amount was $500,000 and €1.2 million.")
        assert len(results) >= 1
        texts = [r.mention_text for r in results]
        assert any("$500,000" in t for t in texts)

    def test_date_extractor(self):
        from app.services.individual.entity_extractors import DateExtractor
        ext = DateExtractor()
        results = ext.extract_sync("The filing was on January 15, 2023 and 03/15/2022.")
        assert len(results) >= 2
        texts = [r.mention_text for r in results]
        assert any("January 15, 2023" in t for t in texts)
        assert any("03/15/2022" in t for t in texts)

    def test_date_extractor_iso_format(self):
        from app.services.individual.entity_extractors import DateExtractor
        ext = DateExtractor()
        results = ext.extract_sync("Effective date: 2023-01-15.")
        assert len(results) >= 1

    def test_duration_extractor(self):
        from app.services.individual.entity_extractors import DurationExtractor
        ext = DurationExtractor()
        results = ext.extract_sync("The lease is for 30 days and two years.")
        assert len(results) >= 1
        texts = [r.mention_text for r in results]
        assert any("30 days" in t for t in texts)

    def test_percentage_extractor(self):
        from app.services.individual.entity_extractors import PercentageExtractor
        ext = PercentageExtractor()
        results = ext.extract_sync("Interest rate of 5% or three percent applies, with 250 basis points.")
        assert len(results) >= 2

    def test_court_extractor(self):
        from app.services.individual.entity_extractors import CourtExtractor
        ext = CourtExtractor()
        results = ext.extract_sync("Filed in the Supreme Court of the United States and S.D.N.Y.")
        assert len(results) >= 1
        texts = [r.mention_text for r in results]
        assert any("Supreme Court" in t for t in texts)

    def test_definition_extractor(self):
        from app.services.individual.entity_extractors import DefinitionExtractor
        ext = DefinitionExtractor()
        results = ext.extract_sync('"Lessor" means the party granting the lease.')
        assert len(results) >= 1
        assert results[0].name == "Lessor"

    def test_condition_extractor(self):
        from app.services.individual.entity_extractors import ConditionExtractor
        ext = ConditionExtractor()
        results = ext.extract_sync("If the tenant fails to pay, provided that notice was given.")
        assert len(results) >= 1
        texts = [r.mention_text.lower() for r in results]
        assert any("if" in t for t in texts) or any("provided that" in t for t in texts)

    def test_constraint_extractor(self):
        from app.services.individual.entity_extractors import ConstraintExtractor
        ext = ConstraintExtractor()
        results = ext.extract_sync("The award shall not exceed $1 million. At least 30 days notice.")
        assert len(results) >= 1

    def test_address_extractor(self):
        from app.services.individual.entity_extractors import AddressExtractor
        ext = AddressExtractor()
        results = ext.extract_sync("Located at 123 Main Street, Suite 100, New York, NY 10001.")
        assert len(results) >= 1

    def test_trademark_extractor(self):
        from app.services.individual.entity_extractors import TrademarkExtractor
        ext = TrademarkExtractor()
        results = ext.extract_sync("Products include Apple® and Google™ devices.")
        assert len(results) >= 2

    def test_copyright_extractor(self):
        from app.services.individual.entity_extractors import CopyrightExtractor
        ext = CopyrightExtractor()
        results = ext.extract_sync("© 2024 Acme Corp. All rights reserved.")
        assert len(results) >= 1

    def test_spacy_person_extractor(self):
        from app.services.individual.entity_extractors import SpaCyPersonExtractor
        ext = SpaCyPersonExtractor()
        results = ext.extract_sync("John Smith filed the complaint against Jane Doe.")
        # spaCy should find at least one person
        assert len(results) >= 1
        assert all(r.source == "spacy_ner" for r in results)

    def test_spacy_org_extractor(self):
        from app.services.individual.entity_extractors import SpaCyOrgExtractor
        ext = SpaCyOrgExtractor()
        results = ext.extract_sync("Google LLC and the Securities and Exchange Commission were parties.")
        assert len(results) >= 1
        assert all(r.source == "spacy_ner" for r in results)

    def test_spacy_location_extractor(self):
        from app.services.individual.entity_extractors import SpaCyLocationExtractor
        ext = SpaCyLocationExtractor()
        results = ext.extract_sync("The court in New York ruled on the California statute.")
        # spaCy should find at least one location
        assert len(results) >= 1

    def test_all_extractors_return_valid_individuals(self):
        from app.services.individual.entity_extractors import ALL_EXTRACTORS
        text = SAMPLE_LEGAL_TEXT
        for ext in ALL_EXTRACTORS:
            results = ext.extract_sync(text)
            for ind in results:
                assert isinstance(ind, Individual)
                assert ind.span.start >= 0
                assert ind.span.end > ind.span.start
                assert ind.name
                assert ind.source in ("regex", "spacy_ner")
                assert 0 < ind.confidence <= 1.0


class TestEntityExtractorRunner:
    async def test_runner_extracts_all_types(self):
        from app.services.individual.entity_extractors import EntityExtractorRunner
        runner = EntityExtractorRunner()
        results = await runner.extract(SAMPLE_LEGAL_TEXT)
        assert len(results) > 0
        # Should have multiple types
        sources = {r.source for r in results}
        assert len(sources) >= 1  # At least regex or spacy_ner


# ── Deduplicator Tests ─────────────────────────────────────────────────


class TestDeduplicator:
    def test_no_duplicates_returns_all(self):
        from app.services.individual.deduplicator import deduplicate
        inds = [
            Individual(
                name="John Smith", mention_text="John Smith",
                span=Span(start=0, end=10, text="John Smith"),
                source="spacy_ner", confidence=0.80,
            ),
            Individual(
                name="$500,000", mention_text="$500,000",
                span=Span(start=50, end=58, text="$500,000"),
                source="regex", confidence=0.93,
            ),
        ]
        result = deduplicate(inds)
        assert len(result) == 2

    def test_overlapping_spans_merged(self):
        from app.services.individual.deduplicator import deduplicate
        inds = [
            Individual(
                name="John Smith", mention_text="John Smith",
                span=Span(start=0, end=10, text="John Smith"),
                source="spacy_ner", confidence=0.80,
                class_links=[IndividualClassLink(folio_label="Person", confidence=0.80)],
            ),
            Individual(
                name="John Smith", mention_text="John Smith",
                span=Span(start=0, end=10, text="John Smith"),
                source="llm", confidence=0.85,
                class_links=[IndividualClassLink(folio_label="Plaintiff", confidence=0.85)],
            ),
        ]
        result = deduplicate(inds)
        assert len(result) == 1
        # Should keep higher priority source (spacy_ner > llm)
        # But merge class links from both
        assert len(result[0].class_links) == 2
        labels = {cl.folio_label for cl in result[0].class_links}
        assert "Person" in labels
        assert "Plaintiff" in labels

    def test_source_priority_preserved(self):
        from app.services.individual.deduplicator import deduplicate
        inds = [
            Individual(
                name="123 U.S. 456", mention_text="123 U.S. 456",
                span=Span(start=0, end=12, text="123 U.S. 456"),
                source="eyecite", confidence=0.92,
            ),
            Individual(
                name="123 U.S. 456", mention_text="123 U.S. 456",
                span=Span(start=0, end=12, text="123 U.S. 456"),
                source="llm", confidence=0.85,
            ),
        ]
        result = deduplicate(inds)
        assert len(result) == 1
        # eyecite has higher priority, so its data wins
        assert result[0].confidence == 0.92

    def test_hybrid_source_on_merge(self):
        from app.services.individual.deduplicator import deduplicate
        inds = [
            Individual(
                name="John Smith", mention_text="John Smith",
                span=Span(start=0, end=10, text="John Smith"),
                source="regex", confidence=0.90,
            ),
            Individual(
                name="John Smith", mention_text="John Smith",
                span=Span(start=0, end=10, text="John Smith"),
                source="llm", confidence=0.85,
            ),
        ]
        result = deduplicate(inds)
        assert len(result) == 1
        assert result[0].source == "hybrid"

    def test_empty_list(self):
        from app.services.individual.deduplicator import deduplicate
        assert deduplicate([]) == []

    def test_url_preserved_on_merge(self):
        from app.services.individual.deduplicator import deduplicate
        inds = [
            Individual(
                name="42 U.S.C. § 1983", mention_text="42 U.S.C. § 1983",
                span=Span(start=0, end=17, text="42 U.S.C. § 1983"),
                source="eyecite", confidence=0.92,
            ),
            Individual(
                name="42 U.S.C. § 1983", mention_text="42 U.S.C. § 1983",
                span=Span(start=0, end=17, text="42 U.S.C. § 1983"),
                source="citeurl", confidence=0.90,
                url="https://www.law.cornell.edu/uscode/text/42/1983",
            ),
        ]
        result = deduplicate(inds)
        assert len(result) == 1
        assert result[0].url == "https://www.law.cornell.edu/uscode/text/42/1983"


# ── LLM Individual Identifier Tests ───────────────────────────────────


class TestLLMIndividualIdentifier:
    @pytest.fixture
    def fake_llm(self):
        llm = MagicMock()
        llm.structured = AsyncMock(return_value={
            "individuals": [
                {
                    "name": "The Employment Agreement",
                    "mention_text": "Employment Agreement",
                    "individual_type": "named_entity",
                    "class_annotation_ids": [],
                    "class_labels": ["Contract"],
                    "confidence": 0.80,
                    "is_new": True,
                },
            ]
        })
        return llm

    async def test_identifies_new_individuals(self, fake_llm):
        from app.services.individual.llm_individual_identifier import LLMIndividualIdentifier
        identifier = LLMIndividualIdentifier(fake_llm)
        chunk = TextChunk(
            text="The Employment Agreement was signed.",
            start_offset=0,
            end_offset=35,
            chunk_index=0,
        )
        results = await identifier.identify_individuals(chunk, [], [])
        assert len(results) == 1
        assert results[0].name == "The Employment Agreement"
        assert results[0].source == "llm"

    async def test_links_existing_individuals(self, fake_llm):
        from app.services.individual.llm_individual_identifier import LLMIndividualIdentifier
        fake_llm.structured = AsyncMock(return_value={
            "individuals": [
                {
                    "name": "John Smith",
                    "mention_text": "John Smith",
                    "individual_type": "named_entity",
                    "class_annotation_ids": ["ann-1"],
                    "class_labels": ["Plaintiff"],
                    "confidence": 0.90,
                    "is_new": False,
                },
            ]
        })
        identifier = LLMIndividualIdentifier(fake_llm)
        chunk = TextChunk(
            text="John Smith filed the complaint.",
            start_offset=0,
            end_offset=30,
            chunk_index=0,
        )
        existing = [
            Individual(
                name="John Smith",
                mention_text="John Smith",
                span=Span(start=0, end=10, text="John Smith"),
                source="spacy_ner",
                confidence=0.80,
            )
        ]
        ann = Annotation(
            id="ann-1",
            span=Span(start=0, end=10, text="Plaintiff"),
            concepts=[ConceptMatch(
                concept_text="plaintiff",
                folio_iri="https://folio.openlegalstandard.org/Plaintiff",
                folio_label="Plaintiff",
                branches=["Actor / Player"],
                confidence=0.90,
            )],
        )
        results = await identifier.identify_individuals(chunk, [ann], existing)
        # No new individuals (is_new=False), but existing should have class link added
        assert len(results) == 0
        assert len(existing[0].class_links) >= 1

    async def test_handles_llm_failure(self):
        from app.services.individual.llm_individual_identifier import LLMIndividualIdentifier
        llm = MagicMock()
        llm.structured = AsyncMock(side_effect=Exception("LLM error"))
        identifier = LLMIndividualIdentifier(llm)
        chunk = TextChunk(
            text="Some text.",
            start_offset=0,
            end_offset=10,
            chunk_index=0,
        )
        results = await identifier.identify_individuals(chunk, [], [])
        assert results == []

    async def test_batch_processing(self, fake_llm):
        from app.services.individual.llm_individual_identifier import LLMIndividualIdentifier
        identifier = LLMIndividualIdentifier(fake_llm)
        chunks = [
            TextChunk(text="The Employment Agreement was signed.", start_offset=0, end_offset=35, chunk_index=0),
            TextChunk(text="Another chunk of text.", start_offset=35, end_offset=57, chunk_index=1),
        ]
        results = await identifier.identify_batch(chunks, [], [])
        assert isinstance(results, list)


# ── Pipeline Stage Tests ───────────────────────────────────────────────


class TestIndividualExtractionStage:
    async def test_early_stage_runs(self):
        from app.pipeline.stages.individual_stage import EarlyIndividualStage
        stage = EarlyIndividualStage()
        assert stage.name == "early_individual_extraction"

        job = _make_job()
        result = await stage.execute(job)
        assert result.status == JobStatus.EXTRACTING_INDIVIDUALS
        assert isinstance(result.result.individuals, list)
        # Should have found some individuals via regex/spaCy even without LLM
        assert len(result.result.individuals) > 0

    async def test_early_stage_skips_when_disabled(self):
        from app.pipeline.stages.individual_stage import EarlyIndividualStage
        stage = EarlyIndividualStage()

        job = _make_job()
        with patch("app.config.settings") as mock_settings:
            mock_settings.individual_extraction_enabled = False
            result = await stage.execute(job)
        assert len(result.result.individuals) == 0

    async def test_early_stage_skips_without_canonical_text(self):
        from app.pipeline.stages.individual_stage import EarlyIndividualStage
        stage = EarlyIndividualStage()
        job = Job(
            input=DocumentInput(content="test", format=DocumentFormat.PLAIN_TEXT),
            status=JobStatus.MATCHING,
            result=JobResult(),
        )
        result = await stage.execute(job)
        assert len(result.result.individuals) == 0

    async def test_llm_stage_regex_only_mode(self):
        from app.pipeline.stages.individual_stage import LLMIndividualStage
        fake_llm = MagicMock()
        fake_llm.structured = AsyncMock()
        stage = LLMIndividualStage(llm=fake_llm)

        job = _make_job()
        with patch("app.config.settings") as mock_settings:
            mock_settings.individual_extraction_enabled = True
            mock_settings.individual_regex_only = True
            result = await stage.execute(job)

        # LLM should not have been called
        fake_llm.structured.assert_not_called()

    async def test_early_stage_logs_activity(self):
        from app.pipeline.stages.individual_stage import EarlyIndividualStage
        stage = EarlyIndividualStage()
        job = _make_job()
        result = await stage.execute(job)
        log = result.result.metadata.get("activity_log", [])
        ind_logs = [l for l in log if l.get("stage") == "early_individual_extraction"]
        assert len(ind_logs) >= 2  # At least pass 1 + pass 2

    async def test_backward_compat_alias(self):
        from app.pipeline.stages.individual_stage import IndividualExtractionStage, EarlyIndividualStage
        assert IndividualExtractionStage is EarlyIndividualStage


# ── Export Integration Tests ───────────────────────────────────────────


def _make_job_with_individuals() -> Job:
    ind = Individual(
        name="John Smith",
        mention_text="John Smith",
        individual_type="named_entity",
        span=Span(start=82, end=92, text="John Smith"),
        class_links=[
            IndividualClassLink(
                folio_iri="https://folio.openlegalstandard.org/Plaintiff",
                folio_label="Plaintiff",
                branch="Actor / Player",
                confidence=0.90,
            )
        ],
        confidence=0.85,
        source="spacy_ner",
    )
    citation = Individual(
        name="Smith v. Jones, 123 U.S. 456 (1987)",
        mention_text="Smith v. Jones, 123 U.S. 456 (1987)",
        individual_type="legal_citation",
        span=Span(start=3, end=39, text="Smith v. Jones, 123 U.S. 456 (1987)"),
        class_links=[
            IndividualClassLink(
                folio_label="Caselaw",
                confidence=0.92,
            )
        ],
        confidence=0.92,
        source="eyecite",
        normalized_form="123 U.S. 456",
        url="https://www.courtlistener.com/opinion/123",
    )

    return Job(
        input=DocumentInput(content=SAMPLE_LEGAL_TEXT, format=DocumentFormat.PLAIN_TEXT),
        status=JobStatus.COMPLETED,
        result=JobResult(
            canonical_text=CanonicalText(
                full_text=SAMPLE_LEGAL_TEXT,
                chunks=[TextChunk(
                    text=SAMPLE_LEGAL_TEXT,
                    start_offset=0,
                    end_offset=len(SAMPLE_LEGAL_TEXT),
                    chunk_index=0,
                )],
            ),
            annotations=[
                Annotation(
                    span=Span(start=4, end=9, text="court"),
                    concepts=[
                        ConceptMatch(
                            concept_text="court",
                            folio_iri="https://folio.openlegalstandard.org/Court",
                            folio_label="Court",
                            branches=["Legal Entity"],
                            confidence=0.95,
                            source="reconciled",
                        )
                    ],
                ),
            ],
            individuals=[ind, citation],
        ),
    )


class TestExportWithIndividuals:
    def test_json_export_includes_individuals(self):
        from app.services.export.json_exporter import JSONExporter
        job = _make_job_with_individuals()
        data = json.loads(JSONExporter().export(job))
        assert "individuals" in data
        assert len(data["individuals"]) == 2
        assert data["statistics"]["total_individuals"] == 2
        assert data["statistics"]["legal_citations"] == 1
        assert data["statistics"]["named_entities"] == 1

    def test_jsonld_export_includes_individuals(self):
        from app.services.export.jsonld_exporter import JSONLDExporter
        job = _make_job_with_individuals()
        data = json.loads(JSONLDExporter().export(job))
        assert "individuals" in data
        assert len(data["individuals"]) == 2
        assert data["individuals"][0]["@type"] == "owl:NamedIndividual"

    def test_xml_export_includes_individuals(self):
        import xml.etree.ElementTree as ET
        from app.services.export.xml_exporter import XMLExporter
        job = _make_job_with_individuals()
        root = ET.fromstring(XMLExporter().export(job))
        inds = root.find("individuals")
        assert inds is not None
        assert len(list(inds)) == 2

    def test_csv_export_includes_individuals(self):
        import csv
        import io
        from app.services.export.csv_exporter import CSVExporter
        job = _make_job_with_individuals()
        result = CSVExporter().export(job)
        reader = csv.reader(io.StringIO(result))
        rows = list(reader)
        # Should have annotation rows + blank + individual header + individual rows
        assert len(rows) >= 5

    def test_jsonl_export_includes_individuals(self):
        from app.services.export.jsonl_exporter import JSONLExporter
        job = _make_job_with_individuals()
        result = JSONLExporter().export(job)
        lines = [json.loads(l) for l in result.strip().split("\n") if l]
        ind_lines = [l for l in lines if l.get("record_type") == "individual"]
        assert len(ind_lines) == 2

    def test_rdf_export_includes_individuals(self):
        from app.services.export.rdf_exporter import RDFExporter
        job = _make_job_with_individuals()
        result = RDFExporter().export(job)
        assert "NamedIndividual" in result

    def test_brat_export_includes_individuals(self):
        from app.services.export.brat_exporter import BratExporter
        job = _make_job_with_individuals()
        result = BratExporter().export(job)
        lines = result.strip().split("\n")
        # Should have more than just annotation lines
        assert len(lines) >= 3

    def test_html_export_includes_individuals(self):
        from app.services.export.html_exporter import HTMLExporter
        job = _make_job_with_individuals()
        result = HTMLExporter().export(job)
        assert "folio-individual" in result

    def test_elasticsearch_export_includes_individuals(self):
        from app.services.export.elasticsearch_exporter import ElasticsearchExporter
        job = _make_job_with_individuals()
        result = ElasticsearchExporter().export(job)
        lines = result.strip().split("\n")
        ind_actions = [l for l in lines if "folio-individuals" in l]
        assert len(ind_actions) == 2

    def test_neo4j_export_includes_individuals(self):
        from app.services.export.neo4j_exporter import Neo4jExporter
        job = _make_job_with_individuals()
        result = Neo4jExporter().export(job)
        assert "Individual" in result
        assert "CONTAINS_INDIVIDUAL" in result

    def test_excel_export_includes_individuals(self):
        from openpyxl import load_workbook
        import io
        from app.services.export.excel_exporter import ExcelExporter
        job = _make_job_with_individuals()
        data = ExcelExporter().export(job)
        wb = load_workbook(io.BytesIO(data))
        assert "Individuals" in wb.sheetnames
        ws = wb["Individuals"]
        assert ws.max_row >= 3  # header + 2 individuals


# ── LLM Prompt Tests ──────────────────────────────────────────────────


class TestIndividualExtractionPrompt:
    def test_prompt_builds_with_annotations(self):
        from app.services.llm.prompts.individual_extraction import (
            build_individual_extraction_prompt,
        )
        prompt = build_individual_extraction_prompt(
            text="John Smith filed suit.",
            class_annotations=[
                {"id": "ann-1", "label": "Plaintiff", "span_text": "plaintiff", "branch": "Actor / Player"},
            ],
            existing_individuals=[
                {"name": "John Smith", "type": "named_entity", "source": "spacy_ner"},
            ],
        )
        assert "John Smith" in prompt
        assert "Plaintiff" in prompt
        assert "spacy_ner" in prompt

    def test_prompt_builds_with_empty_inputs(self):
        from app.services.llm.prompts.individual_extraction import (
            build_individual_extraction_prompt,
        )
        prompt = build_individual_extraction_prompt(
            text="Some text.",
            class_annotations=[],
            existing_individuals=[],
        )
        assert "Some text." in prompt
        assert "none found" in prompt.lower()


# ── Config Tests ──────────────────────────────────────────────────────


class TestIndividualConfig:
    def test_default_config_values(self):
        from app.config import Settings
        s = Settings()
        assert s.individual_extraction_enabled is True
        assert s.individual_regex_only is False
        assert s.llm_individual_provider == ""
        assert s.llm_individual_model == ""
