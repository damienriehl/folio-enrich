"""Excel export with branch colors and confidence color-coding."""

from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from app.models.job import Job
from app.services.export.base import ExporterBase


def _hex_to_rgb(hex_color: str) -> str:
    """Convert '#1a5276' to 'FF1A5276' for openpyxl."""
    h = hex_color.lstrip("#")
    return f"FF{h.upper()}"


def _confidence_fill(confidence: float) -> PatternFill | None:
    """Return a fill color based on confidence score."""
    if confidence >= 0.90:
        return PatternFill(start_color="FF228B22", end_color="FF228B22", fill_type="solid")  # green
    elif confidence >= 0.60:
        return PatternFill(start_color="FFFFD700", end_color="FFFFD700", fill_type="solid")  # gold
    elif confidence >= 0.45:
        return PatternFill(start_color="FFFF8C00", end_color="FFFF8C00", fill_type="solid")  # orange
    return None


class ExcelExporter(ExporterBase):
    @property
    def format_name(self) -> str:
        return "excel"

    @property
    def content_type(self) -> str:
        return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    def export(self, job: Job) -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "FOLIO Annotations"

        # Headers
        headers = [
            "Span Start", "Span End", "Span Text",
            "Concept", "FOLIO IRI", "FOLIO Label",
            "Branch", "Branch Color", "Confidence", "Source",
            "Hierarchy Path", "Definition",
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)

        # Data rows
        row_idx = 2
        for ann in job.result.annotations:
            for concept in ann.concepts:
                branch_color = concept.branch_color or ""
                hierarchy = " > ".join(concept.hierarchy_path) if concept.hierarchy_path else ""

                row_data = [
                    ann.span.start,
                    ann.span.end,
                    ann.span.text,
                    concept.concept_text,
                    concept.folio_iri or "",
                    concept.folio_label or "",
                    concept.branches[0] if concept.branches else "",
                    branch_color,
                    round(concept.confidence, 4),
                    concept.source,
                    hierarchy,
                    concept.folio_definition or "",
                ]
                ws.append(row_data)

                # Apply branch color to the Branch cell
                if branch_color:
                    try:
                        branch_cell = ws.cell(row=row_idx, column=7)
                        fill_color = _hex_to_rgb(branch_color)
                        branch_cell.fill = PatternFill(
                            start_color=fill_color, end_color=fill_color, fill_type="solid"
                        )
                        branch_cell.font = Font(color="FFFFFFFF")
                    except Exception:
                        pass

                # Apply confidence color-coding
                conf_fill = _confidence_fill(concept.confidence)
                if conf_fill:
                    conf_cell = ws.cell(row=row_idx, column=9)
                    conf_cell.fill = conf_fill

                row_idx += 1

        # Auto-width columns
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

        # Individuals sheet
        if job.result.individuals:
            ws_ind = wb.create_sheet("Individuals")
            ind_headers = [
                "Span Start", "Span End", "Mention Text",
                "Name", "Individual Type", "Class Labels",
                "Confidence", "Source", "Normalized Form", "URL",
            ]
            ws_ind.append(ind_headers)
            for cell in ws_ind[1]:
                cell.font = Font(bold=True)

            ind_row_idx = 2
            for ind in job.result.individuals:
                class_labels = "; ".join(
                    cl.folio_label or "" for cl in ind.class_links
                )
                ws_ind.append([
                    ind.span.start,
                    ind.span.end,
                    ind.mention_text,
                    ind.name,
                    ind.individual_type,
                    class_labels,
                    round(ind.confidence, 4),
                    ind.source,
                    ind.normalized_form or "",
                    ind.url or "",
                ])

                # Confidence color-coding
                conf_fill = _confidence_fill(ind.confidence)
                if conf_fill:
                    conf_cell = ws_ind.cell(row=ind_row_idx, column=7)
                    conf_cell.fill = conf_fill

                ind_row_idx += 1

            for col in ws_ind.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                ws_ind.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

        # Properties sheet
        if job.result.properties:
            ws_prop = wb.create_sheet("Properties")
            prop_headers = [
                "Span Start", "Span End", "Property Text",
                "FOLIO IRI", "FOLIO Label", "Confidence",
                "Source", "Match Type", "Domain IRIs", "Range IRIs",
            ]
            ws_prop.append(prop_headers)
            for cell in ws_prop[1]:
                cell.font = Font(bold=True)

            prop_row_idx = 2
            for prop in job.result.properties:
                ws_prop.append([
                    prop.span.start,
                    prop.span.end,
                    prop.property_text,
                    prop.folio_iri or "",
                    prop.folio_label or "",
                    round(prop.confidence, 4),
                    prop.source,
                    prop.match_type or "",
                    "; ".join(prop.domain_iris),
                    "; ".join(prop.range_iris),
                ])

                conf_fill = _confidence_fill(prop.confidence)
                if conf_fill:
                    conf_cell = ws_prop.cell(row=prop_row_idx, column=6)
                    conf_cell.fill = conf_fill

                prop_row_idx += 1

            for col in ws_prop.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                ws_prop.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
